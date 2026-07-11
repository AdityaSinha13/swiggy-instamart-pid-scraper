# =============================================================================
#  🛵 SWIGGY INSTAMART BULK PRODUCT SCRAPER — v2.0
#  Standalone Python Script
# =============================================================================
#
#  Install (sirf ek baar terminal mein):
#      pip install selenium webdriver-manager beautifulsoup4 pandas openpyxl
#
#  Run karo:
#      python swiggy_instamart_scraper.py
#
#  Input file format:
#      Excel/CSV with column "Pids" — Swiggy Instamart Product IDs ya full URLs
#
#  Extraction methods (4 layers in order):
#      1. __NEXT_DATA__ JSON  — Swiggy ka React/Next.js data (most complete)
#      2. JSON-LD schema tags
#      3. OG/Meta tags
#      4. Live DOM XPath fallback
#
# =============================================================================

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup
from pathlib import Path
from time import sleep
import pandas as pd
import json as _json, re, time, random


# =============================================================================
#  ██████╗ ██████╗ ███╗   ██╗███████╗██╗ ██████╗
# ██╔════╝██╔═══██╗████╗  ██║██╔════╝██║██╔════╝
# ██║     ██║   ██║██╔██╗ ██║█████╗  ██║██║  ███╗
# ██║     ██║   ██║██║╚██╗██║██╔══╝  ██║██║   ██║
# ╚██████╗╚██████╔╝██║ ╚████║██║     ██║╚██████╔╝
#  ╚═════╝ ╚═════╝ ╚═╝  ╚═══╝╚═╝     ╚═╝ ╚═════╝
#
#  BAS YAHAN EDIT KARO ↓↓↓
# =============================================================================

# ── File Paths ────────────────────────────────────────────────────────────────
INPUT_FILE  = r"C:\Users\800022918\Downloads\Signify Automation tool\SCRAPPER TOOLS\Swiggy Instamart\Swiggy Input.xlsx"
OUTPUT_FILE = r"C:\Users\800022918\Downloads\Signify Automation tool\SCRAPPER TOOLS\Swiggy Instamart\Swiggy Output.xlsx"

# ── Location ──────────────────────────────────────────────────────────────────
# Pincode ya area naam dono chalega  e.g. "122017"  ya  "Gurugram Haryana"
LOCATION = "122017"

# ── Browser ───────────────────────────────────────────────────────────────────
# BUG FIX: Default False — Swiggy headless mode ko aggressively block karta hai!
# True sirf tab karo jab environment headless support kare (e.g. server/CI)
HEADLESS = False   # False = browser dikhega (recommended) | True = background

# ── Speed ─────────────────────────────────────────────────────────────────────
DELAY_SEC = 3.5    # Har product ke baad wait in seconds (3–5 recommended, kam mat karo!)

# ── Field Selector ────────────────────────────────────────────────────────────
#  True  = yeh column output mein chahiye
#  False = yeh column skip karo
FIELDS = {
    "PID"               : True,   # Product ID
    "Product Name"      : True,   # Product ka full naam
    "Brand"             : True,   # Brand / manufacturer
    "Category"          : True,   # Category hierarchy
    "Selling Price (Rs)": True,   # Current selling price ₹
    "MRP (Rs)"          : True,   # Maximum Retail Price ₹
    "Discount %"        : True,   # Discount % off (auto-calculated)
    "Quantity / Size"   : True,   # Pack size / weight / unit
    "Availability"      : True,   # In Stock / Out of Stock
    "Delivery Time"     : False,  # Estimated delivery time
    "Rating"            : False,  # Star rating (1–5)
    "Rating Count"      : False,  # Total ratings count
    "Description"       : False,  # Product description
    "Product Image URL" : False,  # Direct image link
    "Scraping Status"   : True,   # Success / Failed
    "Error Message"     : False,  # Error details (debugging ke liye)
    "Source URL"        : False,  # Swiggy product page URL
}

# =============================================================================
#  YAHAN SE NEECHE KUCH EDIT KARNE KI ZAROORAT NAHI
# =============================================================================


# ── Browser setup ─────────────────────────────────────────────────────────────
# BUG FIX 1 (headless): Default False — Swiggy blocks headless via navigator checks
# BUG FIX 8 (stealth):  Complete navigator masking — webdriver, plugins, languages, chrome obj
def make_driver(headless=False):
    opts = Options()
    if headless:
        opts.add_argument("--headless=new")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_argument("--disable-infobars")
    opts.add_argument("--window-size=1440,900")
    opts.add_argument("--start-maximized")
    opts.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    )
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_experimental_option("useAutomationExtension", False)
    opts.add_experimental_option("prefs", {
        "credentials_enable_service": False,
        "profile.password_manager_enabled": False,
    })
    driver = webdriver.Chrome(
        service=Service(ChromeDriverManager().install()), options=opts
    )
    # BUG FIX 8: Full stealth — mask all automation fingerprints
    driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {"source": """
        Object.defineProperty(navigator, 'webdriver', {get: () => undefined});
        Object.defineProperty(navigator, 'plugins',   {get: () => [1, 2, 3]});
        Object.defineProperty(navigator, 'languages', {get: () => ['en-US', 'en']});
        window.chrome = { runtime: {} };
        const originalQuery = window.navigator.permissions.query;
        window.navigator.permissions.query = (parameters) =>
            parameters.name === 'notifications'
                ? Promise.resolve({ state: Notification.permission })
                : originalQuery(parameters);
    """})
    return driver


# ── Location Setter ───────────────────────────────────────────────────────────
# BUG FIX 3: Added return value (True/False), verification, retry, better XPaths
def set_location(driver, location_text):
    driver.get("https://www.swiggy.com/instamart")
    sleep(5)

    # Try to open location dialog
    click_xpaths = [
        '//div[@data-testid="search-location"]',
        '//div[contains(@class,"nav-location")]',
        '//div[contains(@class,"location-tab")]',
        '//span[contains(@class,"localize")]',
        '//div[contains(@class,"global-nav")]//div[@role="button"][1]',
        '//div[contains(text(),"Add a new address") or contains(text(),"Enter location")]',
    ]
    for xp in click_xpaths:
        try:
            WebDriverWait(driver, 4).until(
                EC.element_to_be_clickable((By.XPATH, xp))
            ).click()
            sleep(2)
            break
        except Exception:
            pass

    # Find & fill the search input
    input_xpaths = [
        '//input[contains(@placeholder,"Search for area") or contains(@placeholder,"search") or contains(@placeholder,"location")]',
        '//input[@id="location-search-input"]',
        '//input[@type="text"][not(@readonly)][1]',
    ]
    for xp in input_xpaths:
        try:
            box = WebDriverWait(driver, 8).until(
                EC.element_to_be_clickable((By.XPATH, xp))
            )
            box.clear()
            box.send_keys(str(location_text))
            sleep(3)

            # Click first dropdown result
            result_xpaths = [
                '//div[contains(@class,"icon-location-marker")]',
                '//li[contains(@class,"location-item")][1]',
                '//div[@role="option"][1]',
                '//ul/li[1]',
                '//div[contains(@class,"PlaceSuggest")][1]',
                '//div[contains(@class,"suggestion")][1]',
            ]
            clicked = False
            for rxp in result_xpaths:
                try:
                    WebDriverWait(driver, 5).until(
                        EC.element_to_be_clickable((By.XPATH, rxp))
                    ).click()
                    sleep(2)
                    clicked = True
                    break
                except Exception:
                    pass
            if not clicked:
                box.send_keys(Keys.RETURN)
                sleep(2)

            # Confirm button if present
            for cxp in [
                '//button/span[contains(text(),"Confirm")]/..',
                '//button[contains(text(),"Confirm location")]',
                '//button[contains(text(),"Confirm")]',
            ]:
                try:
                    WebDriverWait(driver, 5).until(
                        EC.element_to_be_clickable((By.XPATH, cxp))
                    ).click()
                    sleep(2)
                    break
                except Exception:
                    pass

            sleep(3)
            print(f"✅  Location set — {location_text}")
            return True

        except TimeoutException:
            continue

    print("⚠️   Location input box nahi mila — browser mein manually set karo Swiggy pe")
    return False


# ── Utility Helpers ───────────────────────────────────────────────────────────
def _clean_price(text):
    """Extract clean numeric price string from messy text like '₹199', 'Rs. 45.50'"""
    if not text:
        return ""
    c = re.sub(r"[^\d.]", "", str(text).replace(",", ""))
    parts = c.split(".")
    if len(parts) > 2:
        c = parts[0] + "." + "".join(parts[1:])
    return c if c else ""


# BUG FIX 5: Recursion depth limit — prevents stack overflow on deep/circular JSON
def _deep(obj, key, _depth=0):
    """Recursively search for a key in nested dict/list structures."""
    if _depth > 20:
        return None
    if isinstance(obj, dict):
        if key in obj:
            return obj[key]
        for v in obj.values():
            r = _deep(v, key, _depth + 1)
            if r is not None:
                return r
    elif isinstance(obj, list):
        for i in obj:
            r = _deep(i, key, _depth + 1)
            if r is not None:
                return r
    return None


def _safe_text(driver, xpaths):
    """Try multiple XPaths, return first non-empty text found."""
    for xp in xpaths:
        try:
            t = driver.find_element(By.XPATH, xp).text.strip()
            if t:
                return t
        except Exception:
            pass
    return ""


def _safe_attr(driver, xpaths, attr="src"):
    """Try multiple XPaths, return first non-empty attribute value found."""
    for xp in xpaths:
        try:
            v = driver.find_element(By.XPATH, xp).get_attribute(attr)
            if v:
                return v.strip()
        except Exception:
            pass
    return ""


# BUG FIX 8: Scroll page to trigger lazy-loaded content before scraping
def _scroll_and_wait(driver, delay=1.0):
    try:
        driver.execute_script("window.scrollTo(0, 400);")
        sleep(delay)
        driver.execute_script("window.scrollTo(0, 0);")
        sleep(0.5)
    except Exception:
        pass


# ── Extraction Method 1: __NEXT_DATA__ JSON ───────────────────────────────────
# BUG FIX 6: Broader key walking — original pageProps.product path doesn't exist
#            in Swiggy's current Next.js structure. Now tries 10 candidate paths.
def parse_next_data(soup):
    out = {}
    tag = soup.find("script", id="__NEXT_DATA__")
    if not tag:
        return out
    try:
        nd = _json.loads(tag.string or "")
        pp = nd.get("props", {}).get("pageProps", {})

        product = None
        candidate_fns = [
            lambda: pp.get("product"),
            lambda: pp.get("item"),
            lambda: pp.get("pdp"),
            lambda: pp.get("itemDetails"),
            lambda: _deep(pp, "itemDetails"),
            lambda: _deep(pp, "product"),
            lambda: _deep(pp, "catalogItem"),
            lambda: _deep(nd, "item"),
            lambda: ((_deep(pp, "store") or {}).get("catalogItems") or [{}])[0],
            lambda: _deep(pp, "itemData"),
        ]
        for fn in candidate_fns:
            try:
                result = fn()
                if result and isinstance(result, dict) and result.get("name"):
                    product = result
                    break
            except Exception:
                pass

        if product:
            out["Product Name"] = str(
                product.get("name") or product.get("display_name") or product.get("itemName") or ""
            )
            out["Brand"] = str(
                product.get("brand") or product.get("brand_name") or product.get("brandName") or ""
            )
            out["Category"] = str(
                product.get("category") or product.get("category_name") or product.get("categoryName") or ""
            )

            # Handle paise vs rupees: if value is round and > 1000, divide by 100
            def maybe_paise(v):
                try:
                    f = float(v)
                    return str(round(f / 100, 2)) if f > 1000 and f % 100 == 0 else str(f)
                except Exception:
                    return str(v)

            raw_mrp = product.get("mrp") or product.get("market_price") or product.get("marketPrice") or ""
            raw_sp  = product.get("price") or product.get("selling_price") or product.get("finalPrice") or ""
            out["MRP (Rs)"]           = _clean_price(maybe_paise(raw_mrp))
            out["Selling Price (Rs)"] = _clean_price(maybe_paise(raw_sp))
            out["Quantity / Size"] = str(
                product.get("unit") or product.get("quantity") or
                product.get("weight") or product.get("net_quantity") or
                product.get("variantTag") or ""
            )

            instock = product.get("inStock")
            oos     = product.get("out_of_stock") or product.get("is_sold_out") or product.get("isSoldOut")
            out["Availability"] = "Out of Stock" if (instock is False or oos) else "In Stock"

            img = (
                product.get("image_url") or product.get("image") or
                product.get("thumbnail") or product.get("imageUrl") or ""
            )
            if isinstance(img, dict):
                img = img.get("url") or img.get("src") or ""
            out["Product Image URL"] = str(img)

            out["Rating"]       = str(product.get("avg_rating") or product.get("rating") or product.get("avgRating") or "")
            out["Rating Count"] = str(product.get("rating_count") or product.get("review_count") or product.get("ratingCount") or "")
            out["Description"]  = str(product.get("description") or product.get("item_description") or product.get("itemDescription") or "")
            out["Delivery Time"] = str(_deep(nd, "sla") or _deep(nd, "eta") or _deep(nd, "deliveryTime") or "")

    except Exception:
        pass
    return {k: v for k, v in out.items() if v}


# ── Extraction Method 2: JSON-LD Schema ───────────────────────────────────────
def parse_json_ld(soup):
    out = {}
    for tag in soup.find_all("script", type="application/ld+json"):
        try:
            data = _json.loads(tag.string or "")
            if not isinstance(data, dict):
                continue
            t = data.get("@type", "")
            if t not in ("Product", "ItemPage") and "Product" not in str(t):
                continue
            out["Product Name"]       = out.get("Product Name") or data.get("name", "")
            brand = data.get("brand") or {}
            out["Brand"]              = out.get("Brand") or (brand.get("name") if isinstance(brand, dict) else str(brand))
            out["Category"]           = out.get("Category") or data.get("category", "")
            imgs = data.get("image", [])
            if isinstance(imgs, str):
                imgs = [imgs]
            out["Product Image URL"]  = out.get("Product Image URL") or (imgs[0] if imgs else "")
            ar = data.get("aggregateRating") or {}
            out["Rating"]             = out.get("Rating") or str(ar.get("ratingValue", ""))
            out["Rating Count"]       = out.get("Rating Count") or str(ar.get("reviewCount") or ar.get("ratingCount", ""))
            offers = data.get("offers") or {}
            if isinstance(offers, list):
                offers = offers[0] if offers else {}
            out["Selling Price (Rs)"] = out.get("Selling Price (Rs)") or _clean_price(str(offers.get("price", "")))
            avail = offers.get("availability", "")
            if avail:
                out["Availability"] = "In Stock" if "InStock" in avail else "Out of Stock"
            out["Description"]        = out.get("Description") or data.get("description", "")
        except Exception:
            pass
    return {k: v for k, v in out.items() if v}


# ── Extraction Method 3: OG / Meta Tags ───────────────────────────────────────
def parse_meta(soup):
    out = {}

    def m(prop, name=None):
        tag = soup.find("meta", property=prop) or (
            soup.find("meta", attrs={"name": name}) if name else None
        )
        return (tag.get("content") or "").strip() if tag else ""

    title = m("og:title") or m(None, "title")
    if title:
        out["Product Name"] = re.split(
            r"\s*[|\-–—]\s*(Swiggy|Instamart)", title, flags=re.I
        )[0].strip()
    out["Product Image URL"] = m("og:image")
    desc = m("og:description") or m(None, "description")
    if desc:
        out["Description"] = desc
        pm = re.search(r"₹\s*([\d,]+(?:\.\d+)?)", desc)
        if pm:
            out["Selling Price (Rs)"] = pm.group(1).replace(",", "")
    return {k: v for k, v in out.items() if v}


# ── Extraction Method 4: Live DOM XPath Fallback ──────────────────────────────
def parse_dom(driver):
    out = {}
    out["Product Name"] = _safe_text(driver, [
        "//h1",
        '//*[@data-testid="product_name"]',
        '//*[contains(@class,"ProductName") or contains(@class,"product-name") or contains(@class,"itemName")]',
    ])
    out["Selling Price (Rs)"] = _clean_price(_safe_text(driver, [
        '//*[@data-testid="product_price"]',
        '//*[contains(@class,"finalPrice") or contains(@class,"selling-price")]',
        '//*[contains(@class,"Price") and not(contains(@class,"Mrp")) and not(contains(@style,"line-through"))]',
        '//span[contains(text(),"₹")][not(ancestor::*[contains(@style,"line-through")])][1]',
    ]))
    out["MRP (Rs)"] = _clean_price(_safe_text(driver, [
        '//*[@data-testid="product_mrp"]',
        '//*[contains(@class,"Mrp") or contains(@class,"mrp")]',
        '//*[contains(@style,"line-through")]',
        '//span[contains(@class,"strike") or contains(@class,"crossed")]',
    ]))
    out["Quantity / Size"] = _safe_text(driver, [
        '//*[@data-testid="product_weight"]',
        '//*[contains(@class,"Weight") or contains(@class,"weight")]',
        '//*[contains(@class,"Quantity") or contains(@class,"variantTag")]',
    ])
    out["Brand"] = _safe_text(driver, [
        '//*[contains(@class,"BrandName") or contains(@class,"brand-name") or contains(@class,"brandName")]',
        '//a[contains(@href,"/brand")]',
        '//*[@data-testid="brand_name"]',
    ])
    out["Description"] = _safe_text(driver, [
        '//*[@data-testid="product_description"]',
        '//*[contains(@class,"Description") or contains(@class,"description")]',
    ])
    oos_text = _safe_text(driver, [
        '//*[contains(@class,"SoldOut") or contains(@class,"sold-out") or contains(@class,"outOfStock")]',
        '//button[contains(text(),"Sold") or contains(text(),"Out of Stock") or contains(text(),"Notify")]',
    ])
    out["Availability"] = "Out of Stock" if oos_text else "In Stock"
    out["Delivery Time"] = _safe_text(driver, [
        '//*[contains(@class,"sla") or contains(@class,"delivery-time") or contains(@class,"Eta")]',
        '//span[contains(text(),"min") and not(contains(text(),"ago"))]',
    ])
    out["Rating"]       = _safe_text(driver, ['//*[contains(@class,"Rating")]//span[1]'])
    out["Rating Count"] = _safe_text(driver, ['//*[contains(@class,"RatingCount") or contains(@class,"review-count")]'])
    try:
        crumbs = driver.find_elements(
            By.XPATH, '//nav[contains(@class,"breadcrumb")]//a | //ol//li//a'
        )
        if crumbs:
            out["Category"] = " > ".join(c.text.strip() for c in crumbs if c.text.strip())
    except Exception:
        pass
    out["Product Image URL"] = _safe_attr(driver, [
        '//*[@data-testid="product_image"]//img',
        '//*[contains(@class,"ProductImage") or contains(@class,"product-image")]//img',
        '//main//img[not(contains(@src,"svg"))][1]',
    ], "src")
    return {k: v for k, v in out.items() if v}


# ── Master Extractor ──────────────────────────────────────────────────────────
def extract_product(driver, pid):
    """Run all 4 extraction methods in order, filling gaps. Returns full data dict."""
    result = {
        "PID": pid, "Product Name": "", "Brand": "", "Category": "",
        "Selling Price (Rs)": "", "MRP (Rs)": "", "Discount %": "",
        "Quantity / Size": "", "Availability": "", "Delivery Time": "",
        "Rating": "", "Rating Count": "", "Description": "",
        "Product Image URL": "", "Scraping Status": "Failed",
        "Error Message": "", "Source URL": driver.current_url,
    }
    try:
        _scroll_and_wait(driver, delay=1.0)  # BUG FIX 8: trigger lazy-load
        soup = BeautifulSoup(driver.page_source, "html.parser")

        # Method 1 & 2 — structured data (most reliable)
        for fn in [parse_next_data, parse_json_ld]:
            d = fn(soup)
            result.update({k: v for k, v in d.items() if v and not result.get(k)})

        # Method 3 — meta tags (fill remaining gaps)
        if not result.get("Product Name"):
            d = parse_meta(soup)
            result.update({k: v for k, v in d.items() if v and not result.get(k)})

        # Method 4 — live DOM (last resort)
        if not result.get("Product Name") or not result.get("Selling Price (Rs)"):
            d = parse_dom(driver)
            result.update({k: v for k, v in d.items() if v and not result.get(k)})

        # Auto-calculate discount % if we have both prices
        if not result.get("Discount %") and result.get("MRP (Rs)") and result.get("Selling Price (Rs)"):
            try:
                mrp = float(result["MRP (Rs)"])
                sp  = float(result["Selling Price (Rs)"])
                if mrp > 0 and sp > 0 and mrp >= sp:
                    result["Discount %"] = str(round((mrp - sp) / mrp * 100, 1)) + "%"
            except Exception:
                pass

        result["Scraping Status"] = (
            "Success"
            if (result.get("Product Name") or result.get("Selling Price (Rs)"))
            else "Failed"
        )
        if result["Scraping Status"] == "Failed":
            result["Error Message"] = "No data found — check PID/URL or Swiggy page structure"

    except Exception as e:
        result["Scraping Status"] = "Failed"
        result["Error Message"]   = str(e)

    return result


# ── File Loader ───────────────────────────────────────────────────────────────
# BUG FIX 4 & 9: Fixed URL parsing for Instamart (old code used /prid/ — Swiggy Food pattern!)
# Supports:
#   1. Bare PIDs:         12345  /  ABC-xyz
#   2. Instamart URLs:    https://www.swiggy.com/instamart/item/dettol-hand-wash--12345
#   3. itemId param URLs: ...?itemId=12345
#   4. Old food URLs:     .../prid/12345  (kept for backwards compat)
def load_pids(path):
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"File nahi mila: {path}")
    df = (
        pd.read_csv(p) if p.suffix.lower() in (".csv", ".txt") else pd.read_excel(p)
    )
    col = next(
        (c for c in df.columns if c.strip().lower() in ("pids", "pid")),
        df.columns[0],
    )
    pids = []
    for v in df[col].dropna().astype(str):
        v = v.strip()
        if not v:
            continue
        # BUG FIX: Instamart URL format — /item/{slug}--{id}
        m = re.search(r"/item/[^?/]*--([a-zA-Z0-9]+)", v)
        if m:
            pids.append(m.group(1)); continue
        # Old Swiggy food URL
        m = re.search(r"/prid/([a-zA-Z0-9\-]+)", v)
        if m:
            pids.append(m.group(1)); continue
        # itemId query param
        m = re.search(r"itemId=([a-zA-Z0-9\-]+)", v)
        if m:
            pids.append(m.group(1)); continue
        # Full Instamart URL without -- separator (use last path segment)
        m = re.search(r"swiggy\.com/instamart/(?:item|product)/([a-zA-Z0-9\-]+)/?$", v)
        if m:
            pids.append(m.group(1)); continue
        # Plain PID
        pids.append(v)
    return pids


ALL_COLS = [
    "PID", "Product Name", "Brand", "Category",
    "Selling Price (Rs)", "MRP (Rs)", "Discount %",
    "Quantity / Size", "Availability", "Delivery Time",
    "Rating", "Rating Count", "Description",
    "Product Image URL", "Scraping Status", "Error Message", "Source URL",
]


# BUG FIX 10: get_column_letter() used properly + Swiggy-orange header styling + freeze pane
def save_excel(rows, path, selected_fields):
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)
    cols = [c for c in ALL_COLS if c in selected_fields]
    df = pd.DataFrame(
        [{c: r.get(c, "") for c in cols} for r in rows], columns=cols
    )
    if p.suffix.lower() == ".csv":
        df.to_csv(p, index=False, encoding="utf-8-sig")
        print(f"💾  Saved → {p}")
    else:
        with pd.ExcelWriter(p, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Swiggy Instamart Data")
            ws = writer.sheets["Swiggy Instamart Data"]
            # Swiggy orange header
            header_fill = PatternFill("solid", fgColor="FF6B35")
            for cell in ws[1]:
                cell.font      = Font(bold=True, color="FFFFFF")
                cell.fill      = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
            # BUG FIX 10: Proper column width using get_column_letter
            for col_idx, col_cells in enumerate(ws.columns, 1):
                max_len = max(
                    (len(str(cell.value)) if cell.value is not None else 0)
                    for cell in col_cells
                )
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 60)
            ws.freeze_panes = "A2"
        print(f"💾  Saved → {p}")


# =============================================================================
#  MAIN
# =============================================================================
def main():
    selected = [f for f, v in FIELDS.items() if v]

    print("=" * 65)
    print("  🛵  Swiggy Instamart Bulk Scraper v2.0")
    print("=" * 65)
    print(f"📋  Fields     : {', '.join(selected)}")
    print(f"📍  Location   : {LOCATION}")
    print(f"🖥️   Browser    : {'Headless ⚠️ (may get blocked)' if HEADLESS else 'Visible ✅ (recommended)'}")
    print(f"⏱️   Delay/item : {DELAY_SEC}s")
    print()

    if not selected:
        print("❌  Koi field select nahi! FIELDS dict mein True karo.")
        return

    try:
        pids = load_pids(INPUT_FILE)
        print(f"📥  {len(pids)} PIDs loaded from: {INPUT_FILE}\n")
    except Exception as e:
        print(f"❌  Input file error: {e}")
        return

    if not pids:
        print("❌  Input file mein koi PID nahi mili.")
        return

    driver = None
    results = []
    t0 = time.time()

    try:
        print("🌐  Browser launch ho raha hai...")
        driver = make_driver(headless=HEADLESS)
        print("✅  Browser ready.\n")

        print(f"📍  Location set ho raha hai ({LOCATION})...")
        loc_ok = set_location(driver, LOCATION)
        if not loc_ok:
            print("⚠️   Location set nahi hua.")
            print("     Browser mein Swiggy Instamart pe manually location set karo.")
            input("     ▶️  Set karne ke baad Enter dabao: ")
        print()

        for idx, pid in enumerate(pids, 1):
            print(f"[{idx:>3}/{len(pids)}]  {str(pid)[:25]:<25}  ", end="", flush=True)
            try:
                # BUG FIX 1: Correct Instamart URL format
                url = f"https://www.swiggy.com/instamart/item/{pid}"
                driver.get(url)
                try:
                    WebDriverWait(driver, 15).until(
                        EC.presence_of_element_located((By.XPATH, "//h1 | //main"))
                    )
                except TimeoutException:
                    pass
                sleep(DELAY_SEC + random.uniform(0.5, 1.5))

                # BUG FIX 7: Improved redirect detection (not just exact URL match)
                current  = driver.current_url
                page_src = driver.page_source
                redirected = (
                    "/instamart/item/"    not in current
                    and "/instamart/product/" not in current
                    and "instamart" in current
                )
                if redirected or len(page_src) < 5000:
                    url2 = f"https://www.swiggy.com/instamart/product-detail?itemId={pid}"
                    driver.get(url2)
                    sleep(DELAY_SEC)

                data = extract_product(driver, pid)
                data["Source URL"] = driver.current_url
                results.append(data)

                icon  = "✓" if data["Scraping Status"] == "Success" else "✗"
                name  = (data.get("Product Name") or "N/A")[:35]
                price = data.get("Selling Price (Rs)", "—")
                mrp   = data.get("MRP (Rs)", "—")
                disc  = data.get("Discount %", "")
                avail = data.get("Availability", "")
                print(f"{icon}  {name:<35}  ₹{price}  MRP:₹{mrp}  {disc:>6}  {avail}")

            except Exception as e:
                err = str(e)[:80]
                print(f"✗  Error: {err}")
                results.append({
                    "PID": pid,
                    "Scraping Status": "Failed",
                    "Error Message": err,
                    "Source URL": driver.current_url if driver else "",
                })

            # Checkpoint save every 10 products
            if idx % 10 == 0:
                save_excel(results, OUTPUT_FILE, selected)
                succ_now = sum(1 for r in results if r.get("Scraping Status") == "Success")
                print(f"\n   💾  Checkpoint [{idx}/{len(pids)}] — {succ_now} success so far\n")

            if idx < len(pids):
                sleep(DELAY_SEC + random.uniform(0, 1.2))

    except KeyboardInterrupt:
        print("\n⚠️   User ne stop kiya — partial results save ho rahe hain...")
    finally:
        if driver:
            driver.quit()
            print("\n🔒  Browser band ho gaya.")

    # Final save
    save_excel(results, OUTPUT_FILE, selected)
    elapsed = time.time() - t0
    succ = sum(1 for r in results if r.get("Scraping Status") == "Success")
    fail = len(results) - succ

    print(f"\n{'═' * 65}")
    print(f"✅  Scraping complete!")
    print(f"⏱️   Total time   : {elapsed:.1f}s  ({elapsed / max(len(results), 1):.1f}s per product)")
    print(f"✓   Success      : {succ}")
    print(f"✗   Failed       : {fail}")
    print(f"📊  Total         : {len(results)}")
    print(f"📁  Output file  : {OUTPUT_FILE}")
    if fail > 0:
        print(f"\n💡  Tip: Failed PIDs ke liye — Swiggy pe manually URL khol ke verify karo.")
    print(f"{'═' * 65}")


if __name__ == "__main__":
    main()
